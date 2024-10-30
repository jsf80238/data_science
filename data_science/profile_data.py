import argparse
from collections import Counter, defaultdict
import csv
from datetime import datetime
from pathlib import Path
import os
import random
import re
import shutil
from statistics import mean, quantiles, stdev
import sys
import tempfile
# Imports above are standard Python
# Imports below are 3rd-party
from argparse_range import range_action
import dateutil.parser
from dotenv import dotenv_values
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, borders
from matplotlib import pyplot as plt
import numpy as np
import pandas as pd
import pyarrow.parquet as pq
import seaborn as sns
from pandas import Series

# Imports below are custom
from lib.base import C, Database, Logger, get_line_count

# Excel limitation
MAX_SHEET_NAME_LENGTH = 31
# Excel output
ROUNDING = 1  # 5.4% for example
DATE_FORMAT = "%Y-%m-%dT%H:%M:%S"
# Headings for Excel output
VALUE, COUNT = "Value", "Count"
# Datatypes
NUMBER, DATETIME, STRING = "NUMBER", "DATETIME", "STRING"
# When producing a list of detail values and their frequency of occurrence
DEFAULT_MAX_DETAIL_VALUES = 35
DETAIL_ABBR = " det"
# When analyzing the patterns of a string column
DEFAULT_LONGEST_LONGEST = 50  # Don't display more than this
DEFAULT_MAX_PATTERN_LENGTH = 50
PATTERN_ABBR = " pat"
# Don't plot histograms/boxes if there are fewer than this number of distinct values
# And don't make pie charts if there are more than this number of distinct values
DEFAULT_PLOT_VALUES_LIMIT = 8
# When determining whether a string column could be considered datetime or numeric examine (up to) this number of records
DEFAULT_OBJECT_SAMPLING_COUNT = 500
DEFAULT_OBJECT_CONVERSION_ALLOWED_ERROR_RATE = 5  # %
# Plotting visual effects
PLOT_SIZE_X, PLOT_SIZE_Y = 11, 8.5
PLOT_FONT_SCALE = 0.75
HISTOGRAM, BOX, PIE = "histogram", "box", "pie"
# Good character for spreadsheet-embedded histograms U+25A0
BLACK_SQUARE = "■"
# Output can be to Excel or HTML
EXCEL = "EXCEL"
HTML = "HTML"
OPEN, CLOSE = "{", "}"
FILE_BASE_NAME = "profiled_data"

DATATYPE_MAPPING_DICT = {
    "BIGINT": NUMBER,
    "BINARY": NUMBER,
    "BIT": NUMBER,
    "BOOLEAN": NUMBER,
    "DECIMAL": NUMBER,
    "DOUBLE": NUMBER,
    "FLOAT": NUMBER,
    "INTEGER": NUMBER,
    "NUMERIC": NUMBER,
    "REAL": NUMBER,
    "SMALLINT": NUMBER,
    "TINYINT": NUMBER,
    "VARBINARY": NUMBER,

    "DATE": DATETIME,
    "TIMESTAMP": DATETIME,

    "BLOB": STRING,
    "CHAR": STRING,
    "CLOB": STRING,
    "LONGNVARCHAR": STRING,
    "LONGVARBINARY": STRING,
    "LONGVARCHAR": STRING,
    "NCHAR": STRING,
    "NCLOB": STRING,
    "NVARCHAR": STRING,
    "OTHER": STRING,
    "SQLXML": STRING,
    "TIME": STRING,
    "VARCHAR": STRING,

    # See https://docs.snowflake.com/en/developer-guide/python-connector/python-connector-api#label-python-connector-type-codes
    "0": NUMBER,
    "1": NUMBER,
    "2": STRING,
    "3": DATETIME,
    "4": DATETIME,
    "5": STRING,
    "6": DATETIME,
    "7": DATETIME,
    "8": DATETIME,
    "9": STRING,
    "10": None,
    "11": None,
    "12": STRING,
    "13": NUMBER,
    "14": None,
    "15": None,
    "16": None,
}

ROW_COUNT = "count"
NULL_COUNT = "null"
NULL_PERCENT = "%null"
UNIQUE_COUNT = "unique"
UNIQUE_PERCENT = "%unique"
MOST_COMMON = "most_common"
MOST_COMMON_PERCENT = "%most_common"
LARGEST = "largest"
SMALLEST = "smallest"
LONGEST = "longest"
SHORTEST = "shortest"
MEAN = "mean"
PERCENTILE_25TH = "percentile_25th"
MEDIAN = "median"
PERCENTILE_75TH = "percentile_75th"
STDDEV = "stddev"
FLOAT = "float"

ANALYSIS_LIST = (
    ROW_COUNT,
    NULL_COUNT,
    NULL_PERCENT,
    UNIQUE_COUNT,
    UNIQUE_PERCENT,
    MOST_COMMON,
    MOST_COMMON_PERCENT,
    LARGEST,
    SMALLEST,
    LONGEST,
    SHORTEST,
    MEAN,
    PERCENTILE_25TH,
    MEDIAN,
    PERCENTILE_75TH,
    STDDEV,
)


def format_long_string(s: str, cutoff: int) -> str:
    """
    | This string is really long and won't display nicely ... so adjust, for example:
    |   I'm unhappy with my collection of boring clothes. In striving to cut back on wasteful purchases, and to keep a tighter closet, I have sucked all of the fun out of my closet.
    | Will be replaced with:
    | I'm u...(actual length is 173 characters)...oset.

    :param s: the long string
    :param cutoff: how long is too long
    :return: formatted string
    """
    if not s:
        return ""
    if not isinstance(s, str):
        s = str(s)
    if len(s) <= cutoff:
        return s
    placeholder = f"...(actual length is {len(s)} characters)..."
    prefix = s[:5]
    suffix = s[-5:]
    return prefix + placeholder + suffix


def convert_datatype(name: np.dtypes) -> str:
    """
    Convert Pandas datatypes to a more generic type

    :param value: the Pandas name
    :return: one of: NUMBER, DATETIME, STRING
    """
    name = str(name)
    if "date" in name.lower():
        return DATETIME
    if "int" in name.lower():
        return NUMBER
    if FLOAT in name.lower():
        return NUMBER
    return STRING


def convert_str_to_float(value: str) -> float:
    """
    Convert CSV strings to a useful data type.

    :param value: the value from the CSV column
    :return: the data converted to float
    """
    if value:
        return float(value)
    else:
        return None


def make_sheet_name(s: str, max_length: int, filler: str = "...") -> str:
    """
    For example, make_sheet_name("Hello world!", 7) returns:
    "Hell..."
    """
    # Remove []:*?/\ as these are not valid for sheet names
    illegal_chars = "[]:*?/\\"
    translation_map = str.maketrans(illegal_chars, "_"*len(illegal_chars))
    s = s.translate(translation_map)
    excess_count = len(s) - max_length
    if excess_count <= 0:
        return s
    else:
        return s[:max_length - len(filler)] + filler


def get_pattern(l: list) -> dict:
    """
    | Return a Counter where the keys are the observed patterns and the values are how often they appear.
    |
    | Examples:
    | "hi joe." --> "CC_C(3)"
    | "hello4abigail" --> "C(5)9C(7)"
    | "this+unexpected-   9" --> "C(4)?C(10)?_(3)9"

    :param l: a list of strings
    :return: a pattern analysis
    """
    counter = Counter()
    for value in l:
        if not value or len(value) > max_pattern_length:
            continue
        value = re.sub("[a-zA-Z]", "C", value)  # Replace letters with 'C'
        value = re.sub(r"\d", "9", value)  # Replace numbers with '9'
        value = re.sub(r"\s+", "_", value)  # Replace whitespace with '_'
        value = re.sub(r"\W", "?", value)  # Replace anything else with '?'
        # Group long sequences of letters or numbers
        # See https://stackoverflow.com/questions/76230795/replace-characters-with-a-count-of-characters
        # The number below (2) means sequences of 3 or more will be grouped
        value = re.sub(r'(.)\1{2,}', lambda m: f'{m.group(1)}({len(m.group())})', value)
        counter[value] += 1
    return counter


def make_html_header(title: str, root_output_file: str = None) -> str:
    """
    | Creates the first part of an HTML file.
    | Used when creating HTML output.

    :param title: displayed by the browser
    :param root_output_file: path to home page, None if creating the home page
    :return: <!DOCTYPE html> <html lang="en-US"> <head> ...
    """
    if root_output_file:
        home_link = f'<a href="../{root_output_file.name}.html">Home</a>'
    else:
        home_link = ""
    if input_query:
        title = input_query
    elif input_path:
        title = input_path.name
    else:
        raise Exception("Programming error")
    return f"""
    <!DOCTYPE html>
    <html lang="en-US">
        <head>
        <meta charset="utf-8">
        <title>Exploratory Data Analysis</title>
        <style>
            html {OPEN}
                font-family: monospace;
                font-size: smaller;
            {CLOSE}
            h1, h2, h3, p, img, div {OPEN}
                text-align: center;
            {CLOSE}
            table, tr, td {OPEN}
                border: 1px solid #000;
                margin-left: auto;
                margin-right: auto;
            {CLOSE}
        </style>
        </head>
        <body>
            <h1>Exploratory Data Analysis for {title}</h1>
            <p>{home_link}</p>
            <div>
    """


def make_html_footer() -> str:
    """
    | Helper function

    :return: </body> </html>
    """
    return f"""
            </div>
            </body>
        </html>
    """


def is_data_file(input_argument: str) -> bool:
    """
    Use the text of the input (select statement or file name) to determine if this is a file or a select statement
    :param input_argument: what the user provided
    :return: True if we think this is a data file, else False
    """
    for suffix in C.PARQUET_EXTENSION.value, C.CSV_EXTENSION.value, ".dat", ".txt", ".dsv":
        if input_argument.lower().endswith(suffix):
            return True
    return False


def insert_image(image_type: str, sheet_number: int) -> None:
    """
    Add a image/plot to a new sheet
    :image_type: HISTOGRAM, BOX, or PIE
    :sheet_number: where the image should be added
    """
    target_sheet_name = make_sheet_name(column_name, MAX_SHEET_NAME_LENGTH - 4) + " " + image_type[:3]
    workbook.create_sheet(target_sheet_name, sheet_number)
    worksheet = workbook.worksheets[sheet_number]
    image_path = tempdir_path / (f"{column_name}.{image_type}.png")
    logger.info(f"Adding {image_path} to {output_file} as sheet '{target_sheet_name}' ...")
    image = openpyxl.drawing.image.Image(image_path)
    image.anchor = "A1"
    worksheet.add_image(image)


def assign_best_datatype(s: Series) -> str:
    """
    Pandas will automatically attempt to convert data to datetime where possible
    It does so poorly, in my experience, so will also try "manually"
    For example, https://www.kaggle.com/datasets/philipagbavordoe/car-prices
    """
    # Sample up to DATATYPE_SAMPLING_SIZE non-null values
    a_sample = s.sample(min(object_sampling_limit, s.size))
    failure_count = 0
    for item in list(a_sample):
        try:
            convert_str_to_datetime(item)
        except Exception as e:
            failure_count += 1
    failure_ratio = failure_count / a_sample.size
    if failure_ratio <= object_conversion_allowed_error_rate:
        logger.info(f"Casting column '{column_name}' as a datetime ...")
        s = pd.to_datetime(s.apply(safe_convert_str_to_datetime))
        return DATETIME, s
    else:
        logger.info(
            f"Error rate of {100 * failure_ratio:.1f}% when attempting to cast column '{column_name}' as a datetime.")
        failure_count = 0
        for item in list(a_sample):
            try:
                float(item)
            except Exception as e:
                failure_count += 1
        failure_ratio = failure_count / a_sample.size
        if failure_ratio <= object_conversion_allowed_error_rate:
            logger.info(f"Casting column '{column_name}' as numeric ...")
            s = pd.to_numeric(s, errors="coerce")
            return NUMBER, s
        else:
            logger.info(f"Error rate of {100 * failure_ratio:.1f}% when attempting to cast column '{column_name}' as numeric.")
    logger.info(f"Casting column '{column_name}' as a string ...")
    return STRING, s


def convert_str_to_datetime(value: str) -> pd.Timestamp:
    """
    Pandas uses dateutils to parse dates. And whereas Python supports dates from year 0000 to 9999, Pandas does not.
    :param value: the string which might be a datetime
    :return: the value as a naive datetime (openpyxl does not support timezones)
    """
    nominal_result = pd.to_datetime(value).replace(tzinfo=None)
    if nominal_result > pd.Timestamp.max:
        return pd.Timestamp.max
    return max(pd.Timestamp.min, nominal_result)


def safe_convert_str_to_datetime(value: str) -> pd.Timestamp:
    """
    This is a wrapper around convert_str_to_datetime which swallows exceptions when used with Pandas' apply function
    :param value: the string which might be a datetime
    :return: the value as a naive datetime (openpyxl does not support timezones), or None if it cannot be parsed
    """
    try:
        return convert_str_to_datetime(value)
    except:
        return None


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description='Profile the data in a database or file. Generates an analysis consisting tables and images stored in an Excel workbook or HTML pages. For string columns provides a pattern analysis with C replacing letters, 9 replacing numbers, underscore replacing spaces, and question mark replacing everything else. For numeric and datetime columns produces a histogram and box plots.')

    parser.add_argument('input',
                        metavar="/path/to/input_data_file.extension | query-against-database",
                        help="An example query is 'select a, b, c from t where x>7'. File names must end in csv, dat, txt, dsv or parquet. See also --delimiter.")
    parser.add_argument('--header-lines',
                        type=int,
                        metavar="NUM",
                        action=range_action(1, sys.maxsize),
                        default=0,
                        help="When reading from a file specifies the number of rows to skip UNTIL the header row. Ignored when getting data from a database. Default is 0.")
    parser.add_argument('--delimiter',
                        metavar="CHAR",
                        default=",",
                        help="Use this character to delimit columns, default is a comma. Ignored when getting data from a database or a parquet file.")
    parser.add_argument('--sample-rows',
                        type=int,
                        metavar="NUM",
                        action=range_action(1, sys.maxsize),
                        help=f"When reading from a file randomly choose this number of rows. If greater than or equal to the number of data rows will use all rows. Ignored when getting data from a database.")
    parser.add_argument('--max-detail-values',
                        type=int,
                        metavar="NUM",
                        action=range_action(1, sys.maxsize),
                        default=DEFAULT_MAX_DETAIL_VALUES,
                        help=f"Produce this many of the top value occurrences, default is {DEFAULT_MAX_DETAIL_VALUES}.")
    parser.add_argument('--max-pattern-length',
                        type=int,
                        metavar="NUM",
                        action=range_action(1, sys.maxsize),
                        default=DEFAULT_MAX_PATTERN_LENGTH,
                        help=f"When segregating strings into patterns leave untouched strings of length greater than this, default is {DEFAULT_MAX_PATTERN_LENGTH}.")
    parser.add_argument('--plot-values-limit',
                        type=int,
                        metavar="NUM",
                        action=range_action(1, sys.maxsize),
                        default=DEFAULT_PLOT_VALUES_LIMIT,
                        help=f"Don't make histograms or box plots when there are fewer than this number of distinct values, and don't make pie charts when there are more than this number of distinct values, default is {DEFAULT_PLOT_VALUES_LIMIT}.")
    parser.add_argument('--no-pattern',
                        action='store_true',
                        help=f"Don't identify patterns in text columns.")
    parser.add_argument('--no-histogram',
                        action='store_true',
                        help=f"Don't make histograms.")
    parser.add_argument('--no-box',
                        action='store_true',
                        help=f"Don't make box plots.")
    parser.add_argument('--no-pie',
                        action='store_true',
                        help=f"Don't make pie charts.")
    parser.add_argument('--no-visual',
                        action='store_true',
                        help=f"Don't make histograms or box plots or pie charts.")
    parser.add_argument('--max-longest-string',
                        type=int,
                        metavar="NUM",
                        action=range_action(50, sys.maxsize),
                        default=DEFAULT_LONGEST_LONGEST,
                        help=f"When displaying long strings show a summary if string exceeds this length, default is {DEFAULT_LONGEST_LONGEST}.")
    parser.add_argument('--object-sampling-limit',
                        metavar="NUM",
                        action=range_action(1, sys.maxsize),
                        default=DEFAULT_OBJECT_SAMPLING_COUNT,
                        help=f"To determine whether a string column can be treated as datetime or numeric sample this number of values, default is {DEFAULT_OBJECT_SAMPLING_COUNT}.")
    parser.add_argument('--object-conversion-allowed-error-rate',
                        metavar="NUM",
                        action=range_action(1, 100),
                        default=DEFAULT_OBJECT_CONVERSION_ALLOWED_ERROR_RATE,
                        help=f"To determine whether a string column can be treated as datetime or numeric allow up to this percentage of values to remain un-parseable, default is {DEFAULT_OBJECT_CONVERSION_ALLOWED_ERROR_RATE}.")
    parser.add_argument('--target-dir',
                        metavar="/path/to/dir",
                        default=Path.cwd(),
                        help="Default is the current directory. Will make intermediate directories as necessary.")
    parser.add_argument('--html',
                        action='store_true',
                        help="Also produce a zip file containing the results in HTML format.")
    parser.add_argument('--get-cleaned-version',
                        metavar="FILE_NAME",
                        help=f"Output the Pandas data frame in CSV or Parquet format. Might be useful if string columns were converted to datetimes/numerics. File name must end in '{C.CSV_EXTENSION}' or '{C.PARQUET_EXTENSION}'.")
    parser.add_argument('--db-host-name',
                        metavar="HOST_NAME",
                        help="Overrides HOST_NAME environment variable. Ignored when getting data from a file.")
    parser.add_argument('--db-port-number',
                        metavar="PORT_NUMBER",
                        help="Overrides PORT_NUMBER environment variable. Ignored when getting data from a file.")
    parser.add_argument('--db-name',
                        metavar="DATABASE_NAME",
                        help="Overrides DATABASE_NAME environment variable. Ignored when getting data from a file.")
    parser.add_argument('--db-user-name',
                        metavar="USER_NAME",
                        help="Overrides USER_NAME environment variable. Ignored when getting data from a file.")
    parser.add_argument('--db-password',
                        metavar="PASSWORD",
                        help="Overrides PASSWORD environment variable. Ignored when getting data from a file.")
    parser.add_argument('--environment-file',
                        metavar="/path/to/file",
                        help="An additional source of database connection information. Overrides environment settings.")

    logging_group = parser.add_mutually_exclusive_group()
    logging_group.add_argument('--verbose', action='store_true')
    logging_group.add_argument('--terse', action='store_true')

    args = parser.parse_args()
    if is_data_file(args.input):
        input_path = Path(args.input)
        input_query = ""
    else:
        input_path = ""
        input_query = args.input
    host_name = args.db_host_name
    port_number = args.db_port_number
    database_name = args.db_name
    user_name = args.db_user_name
    password = args.db_password
    if args.environment_file:
        environment_file = Path(args.environment_file)
    else:
        environment_file = ""
    header_lines = args.header_lines
    delimiter = args.delimiter
    sample_rows = args.sample_rows
    max_detail_values = args.max_detail_values
    max_pattern_length = args.max_pattern_length
    max_longest_string = args.max_longest_string
    plot_values_limit = args.plot_values_limit
    object_sampling_limit = args.object_sampling_limit
    object_conversion_allowed_error_rate = args.object_conversion_allowed_error_rate / 100
    cleaned_version_output_file = args.get_cleaned_version
    is_html_output = args.html
    is_excel_output = True
    is_pattern = not args.no_pattern
    is_histogram = not args.no_histogram and not args.no_visual
    is_box = not args.no_box and not args.no_visual
    is_pie = not args.no_pie and not args.no_visual
    target_dir = Path(args.target_dir)

    environment_settings_dict = {
        **os.environ,
        **dotenv_values(environment_file),
    }
    if not target_dir.parent.exists():
        parser.error("Directory '{output_dir.parent}' does not exist.")
    else:
        os.makedirs(target_dir, exist_ok=True)

    if input_query:
        # Verify we have the information we need to connect to the database
        host_name = host_name or environment_settings_dict.get("HOST_NAME")
        if not host_name:
            parser.error("Connecting to a database requires environment variable and/or environment file and/or --db-host-name.")
        port_number = port_number or environment_settings_dict.get("PORT_NUMBER")
        if not port_number:
            parser.error("Connecting to a database requires environment variable and/or environment file and/or --db-port-number.")
        database_name = database_name or environment_settings_dict.get("DATABASE_NAME")
        if not database_name:
            parser.error("Connecting to a database requires environment variable and/or environment file and/or --db-database-name.")
        user_name = user_name or environment_settings_dict.get("USER_NAME")
        if not user_name:
            parser.error("Connecting to a database requires environment variable and/or environment file and/or --db-user-name.")
        password = password or environment_settings_dict.get("PASSWORD")
        if not password:
            parser.error("Connecting to a database requires environment variable and/or environment file and/or --db-password.")
    elif input_path:
        if not input_path.exists():
            parser.error(f"Could not find input file '{input_path}'.")
    else:
        raise Exception("Programming error.")

    if cleaned_version_output_file:
        if not (cleaned_version_output_file.lower().endswith(C.CSV_EXTENSION) or cleaned_version_output_file.lower().endswith(C.PARQUET_EXTENSION)):
            parser.error(f"Cleaned version output file name, if provided, must end with '{C.CSV_EXTENSION}' or '{C.PARQUET_EXTENSION}'.")
    if args.verbose:
        logger = Logger("DEBUG").get_logger()
    elif args.terse:
        logger = Logger("WARNING").get_logger()
    else:
        logger = Logger().get_logger()


    # Verify we have permission to write to the output file
    if is_excel_output:
        output_file = (target_dir / f"{FILE_BASE_NAME}{C.EXCEL_EXTENSION}")
        if output_file.exists():
            os.remove(output_file)
    # Now, read the data
    input_df = None
    data_dict = defaultdict(list)
    # ↑ Keys are column_names, values are a list of values from the data.
    datatype_dict = dict()
    # ↑ Keys are column_names, values are the type of data (NUMBER, DATETIME, STRING)
    if input_query:
        # Data is coming from a database query
        mydb = Database(
            host_name=host_name,
            port_number=port_number,
            database_name=database_name,
            user_name=user_name,
            password=password
        )
        cursor, column_list = mydb.execute(input_query)
        for r in cursor.fetchall():
            row = dict(zip(column_list, r))
            # Store data
            for column_name, value in row.items():
                data_dict[column_name].append(value)
        # Provide row count
        for key, value_list in data_dict.items():
            logger.info(f"Data read: {len(value_list)} rows.")
            break
        # Determine datatype
        for item in cursor.description:
            column_name, dbapi_type_code, display_size, internal_size, precision, scale, null_ok = item
            type_code_desc = str(dbapi_type_code).upper()
            # ↑ Converts things like DBAPITypeObject('BOOLEAN', 'BIGINT', 'BIT', 'INTEGER', 'SMALLINT', 'TINYINT') to
            # "DBAPITypeObject('BOOLEAN', 'BIGINT', 'BIT', 'INTEGER', 'SMALLINT', 'TINYINT')", which is good enough
            # to determine the datatype.
            for key in DATATYPE_MAPPING_DICT:
                if key in type_code_desc:
                    datatype_dict[column_name] = DATATYPE_MAPPING_DICT[key]
                    logger.info(f"Read column '{column_name}' as {DATATYPE_MAPPING_DICT[key]}.")
                    break
            else:
                logger.error(f"Could not determine data type for column '{column_name}' based on the JDBC metadata: {str(item)}")
                datatype_dict[column_name] = STRING
        # Sometimes the JDBC API returns strings for values its metadata says are dates/datetimes.
        # Convert these as necessary from Python strings to Python datetimes
        for column_name, values in data_dict.items():
            if datatype_dict[column_name] == DATETIME:
                # Check the type of the first non-null value
                if isinstance(data_dict[column_name][0], str):
                    data_dict[column_name] = list(map(lambda x: dateutil.parser.parse(x), data_dict[column_name]))
        input_df = pd.DataFrame.from_dict(data_dict)

    elif input_path:
        # Data is coming from a file
        logger.info(f"Reading from '{input_path}' ...")
        if input_path.name.endswith(C.PARQUET_EXTENSION.value):
            input_df = pq.read_table(input_path).to_pandas()
        else:
            input_df = pd.read_csv(input_path, delimiter=delimiter, header=header_lines)
        # Sampling requested?
        if sample_rows:
            input_df = input_df.sample(sample_rows)

    # Data has been read into input_df, now process it
    if input_df.shape[0] == 0:
        logger.critical(f"There is no data in '{input_path + input_query}'.")
        exit()
    """
    Pandas will automatically attempt to convert data to datetime where possible
    It does so poorly, in my experience, so will also try "manually"
    For example, https://www.kaggle.com/datasets/philipagbavordoe/car-prices
    """
    for column_name in input_df.select_dtypes(include=["object"]).columns:
        logger.info(f"Examining datatype for column '{column_name}' ...")
        mask = (input_df[column_name].isna() | input_df[column_name].isnull())
        s = input_df[~mask][column_name]
        if not s.size:
            # Column empty, we don't care about the type
            continue
        # Sample up to DATATYPE_SAMPLING_SIZE non-null values
        a_sample = s.sample(min(object_sampling_limit, s.size))
        failure_count = 0
        for item in list(a_sample):
            try:
                convert_str_to_datetime(item)
            except Exception as e:
                failure_count += 1
        failure_ratio = failure_count / a_sample.size
        if failure_ratio <= object_conversion_allowed_error_rate:
            logger.info(f"Casting column '{column_name}' as a datetime ...")
            input_df[column_name] = pd.to_datetime(s.apply(safe_convert_str_to_datetime))
            datatype_dict[column_name] = DATETIME
        else:
            logger.debug(f"Error rate of {100 * failure_ratio:.1f}% when attempting to cast column '{column_name}' as a datetime.")
            failure_count = 0
            for item in list(a_sample):
                try:
                    float(item)
                except Exception as e:
                    failure_count += 1
            failure_ratio = failure_count / a_sample.size
            if failure_ratio <= object_conversion_allowed_error_rate:
                logger.info(f"Casting column '{column_name}' as numeric ...")
                input_df[column_name] = pd.to_numeric(s, errors="coerce")
                datatype_dict[column_name] = NUMBER
            else:
                logger.debug(f"Error rate of {100 * failure_ratio:.1f}% when attempting to cast column '{column_name}' as numeric.")
                logger.info(f"Will keep column '{column_name}' as a string.")
                datatype_dict[column_name] = STRING
    # To temporarily hold plots and html files
    tempdir = tempfile.TemporaryDirectory()
    tempdir_path = Path(tempdir.name)
    # To keep track of which columns have histogram plots
    histogram_plot_list = list()
    box_plot_list = list()
    pie_plot_list = list()
    # Standardize column names
    input_df.columns = [make_sheet_name(x, MAX_SHEET_NAME_LENGTH) for x in input_df.columns]

    summary_dict = dict()  # To be converted into the summary worksheet
    detail_dict = dict()  # Each element to be converted into a detail worksheet
    pattern_dict = dict()  # For each string column calculate the frequency of patterns
    # for column_name, datatype in dict(input_df.dtypes).items():
    for column_name in input_df.columns:
        values = input_df[column_name]
        if False and not column_name.startswith("pre"):  # For testing
            continue
        # A list of non-null values are useful for some calculations below
        mask = (input_df[column_name].isna() | input_df[column_name].isnull())
        non_null_df = input_df[~mask][column_name].to_frame()
        logger.info(f"Working on column '{column_name}' ...")
        column_dict = dict.fromkeys(ANALYSIS_LIST)
        # Row count
        row_count = values.size
        column_dict[ROW_COUNT] = row_count
        # Null
        null_count = row_count - non_null_df.shape[0]
        column_dict[NULL_COUNT] = null_count
        # Null%
        column_dict[NULL_PERCENT] = round(100 * null_count / row_count, ROUNDING)
        # Unique
        unique_count = values.nunique(dropna=False)
        column_dict[UNIQUE_COUNT] = unique_count
        # Unique%
        column_dict[UNIQUE_PERCENT] = round(100 * unique_count / row_count, ROUNDING)

        # Convert the various types of Pandas datatypes into one of: STRING, NUMBER, DATETIME
        datatype = datatype_dict[column_name]

        if null_count != row_count:

            temp = "temp"
            if datatype == STRING:
                # Largest & smallest
                column_dict[LARGEST] = format_long_string(non_null_df[column_name].max(), max_longest_string)
                column_dict[SMALLEST] = format_long_string(non_null_df[column_name].min(), max_longest_string)
                # Longest & shortest
                non_null_df[temp] = non_null_df[column_name].str.len()
                x = non_null_df.sort_values([temp], ascending=True, axis=0)
                column_dict[SHORTEST] = x[column_name].iloc[0]
                x = non_null_df.sort_values([temp], ascending=False, axis=0)
                longest_string = x[column_name].iloc[0]
                column_dict[LONGEST] = format_long_string(longest_string, max_longest_string)
                # No mean/quartiles/stddev statistics for strings
            elif datatype == NUMBER:
                # Largest & smallest
                column_dict[LARGEST] = non_null_df[column_name].max()
                column_dict[SMALLEST] = non_null_df[column_name].min()
                # No longest/shortest for numbers and dates
                column_dict[SHORTEST] = np.nan
                column_dict[LONGEST] = np.nan
                # Mean/quartiles/stddev statistics
                column_dict[MEAN] = non_null_df[column_name].astype(FLOAT).mean()
                column_dict[STDDEV] = non_null_df[column_name].astype(FLOAT).std()
                column_dict[PERCENTILE_25TH] = non_null_df[column_name].astype(FLOAT).quantile(0.25)
                column_dict[MEDIAN] = non_null_df[column_name].astype(FLOAT).quantile(0.5)
                column_dict[PERCENTILE_75TH] = non_null_df[column_name].astype(FLOAT).quantile(0.75)
            elif datatype == DATETIME:
                # Largest & smallest
                # For the next two lines convert datetime to string because openpyxl does not support datetimes with timezones
                column_dict[LARGEST] = non_null_df[column_name].max().strftime(DATE_FORMAT)
                column_dict[SMALLEST] = non_null_df[column_name].min().strftime(DATE_FORMAT)
                # No longest/shortest for numbers and dates
                column_dict[SHORTEST] = np.nan
                column_dict[LONGEST] = np.nan
                # Mean/quartiles/stddev statistics
                non_null_df[temp] = non_null_df[column_name].astype('int64') // 1e9
                # For the next four lines convert datetime to string because openpyxl does not support datetimes with timezones
                column_dict[MEAN] = datetime.fromtimestamp(non_null_df[temp].mean()).strftime(DATE_FORMAT)
                column_dict[PERCENTILE_25TH] = datetime.fromtimestamp(non_null_df[temp].quantile(0.25)).strftime(DATE_FORMAT)
                column_dict[MEDIAN] = datetime.fromtimestamp(non_null_df[temp].quantile(0.5)).strftime(DATE_FORMAT)
                column_dict[PERCENTILE_75TH] = datetime.fromtimestamp(non_null_df[temp].quantile(0.75)).strftime(DATE_FORMAT)
                column_dict[STDDEV] = non_null_df[temp].std() / 24 / 60 / 60  # Report standard deviation of datetimes in units of days
            else:
                raise Exception("Programming error.")

            summary_dict[column_name] = column_dict

            # Value counts
            # Collect no more than number of values available or what was given on the command-line
            # whichever is less
            counter = Counter(values.to_list())
            max_length = min(max_detail_values, values.size)
            most_common_list = counter.most_common(max_length)
            if datatype == DATETIME:
                most_common_datetime_list = list()
                for item, count in most_common_list:
                    # Convert datetime to string because openpyxl does not support datetimes with timezones
                    try:
                        most_common_datetime_list.append((item.strftime(DATE_FORMAT), count))
                    except ValueError as e:
                        most_common_datetime_list.append(("", count))
                most_common_list = most_common_datetime_list
            most_common, most_common_count = most_common_list[0]
            column_dict[MOST_COMMON] = most_common
            column_dict[MOST_COMMON_PERCENT] = round(100 * most_common_count / row_count, ROUNDING)
            detail_df = pd.DataFrame()
            # Create 3-column descending visual
            detail_df["rank"] = list(range(1, len(most_common_list) + 1))
            detail_df["value"] = [x[0] for x in most_common_list]
            detail_df["count"] = [x[1] for x in most_common_list]
            detail_df["%total"] = [round(x[1] * 100 / row_count, ROUNDING) for x in most_common_list]
            detail_df["histogram"] = [BLACK_SQUARE * round(x[1] * 100 / row_count) for x in most_common_list]
            detail_dict[column_name] = detail_df
        else:
            logger.warning(f"Column '{column_name}' is empty.")

        # Produce visuals
        values = pd.Series(values)
        plot_data = values.value_counts(normalize=True)
        # Produce a pattern analysis for strings
        if is_pattern and datatype == STRING and row_count:
            pattern_counter = get_pattern(non_null_df[column_name])
            max_length = min(max_detail_values, len(non_null_df))
            most_common_pattern_list = pattern_counter.most_common(max_length)
            pattern_df = pd.DataFrame()
            # Create 3-column descending visual
            pattern_df["rank"] = list(range(1, len(most_common_pattern_list) + 1))
            pattern_df["pattern"] = [x[0] for x in most_common_pattern_list]
            pattern_df["count"] = [x[1] for x in most_common_pattern_list]
            pattern_df["%total"] = [round(x[1] * 100 / row_count, ROUNDING) for x in most_common_pattern_list]
            pattern_df["histogram"] = [BLACK_SQUARE * round(x[1] * 100 / row_count) for x in most_common_pattern_list]
            pattern_dict[column_name] = pattern_df
        else:  # Numeric/datetime data
            try:
                values_to_plot = values.astype(FLOAT)
            except TypeError as e:
                values_to_plot = values
            if is_histogram and len(plot_data) >= plot_values_limit:
                logger.info("Creating a histogram plot ...")
                plot_output_path = tempdir_path / f"{column_name}.histogram.png"
                plt.figure(figsize=(PLOT_SIZE_X/2, PLOT_SIZE_Y/2))
                ax = values_to_plot.plot.hist(bins=min(20, len(values_to_plot)))
                ax.set_xlabel(column_name)
                ax.set_ylabel("Count")
                plt.savefig(plot_output_path)
                plt.close('all')  # Save memory
                logger.info(f"Wrote {os.stat(plot_output_path).st_size:,} bytes to '{plot_output_path}'.")
                histogram_plot_list.append(column_name)
            if is_box and len(non_null_df) >= plot_values_limit:
                logger.info("Creating box plots ...")
                plot_output_path = tempdir_path / f"{column_name}.box.png"
                fig, axs = plt.subplots(
                    nrows=2,
                    ncols=1,
                    figsize=(PLOT_SIZE_X, PLOT_SIZE_Y)
                )
                sns.boxplot(
                    ax=axs[0],
                    data=None,
                    x=values_to_plot,
                    showfliers=True,
                    orient="h"
                )
                axs[0].set_xlabel(f"'{column_name}' with outliers")
                sns.boxplot(
                    ax=axs[1],
                    data=None,
                    x=values_to_plot,
                    showfliers=False,
                    orient="h"
                )
                axs[1].set_xlabel(f"'{column_name}' without outliers")
                #plt.subplots_adjust(left=1, right=4, bottom=0.75, top=3, wspace=0.5, hspace=3)
                plt.savefig(plot_output_path)
                plt.close('all')  # Save memory
                logger.info(f"Wrote {os.stat(plot_output_path).st_size:,} bytes to '{plot_output_path}'.")
                box_plot_list.append(column_name)
        if is_pie and len(plot_data) < plot_values_limit:
            logger.info("Creating pie plot ...")
            plot_output_path = tempdir_path / f"{column_name}.pie.png"
            s = values.value_counts()
            fig, ax = plt.subplots()
            ax.pie(s, labels=s.index, autopct="%1.1f%%")
            ax.set_title(column_name)
            plt.savefig(plot_output_path)
            plt.close('all')  # Save memory
            logger.info(f"Wrote {os.stat(plot_output_path).st_size:,} bytes to '{plot_output_path}'.")
            pie_plot_list.append(column_name)

    # Convert the summary_dict dictionary of dictionaries to a DataFrame
    result_df = pd.DataFrame.from_dict(summary_dict, orient='index')

    # Output
    if is_excel_output:
        logger.info("Writing summary ...")
        output_file = (target_dir / f"{FILE_BASE_NAME}{C.EXCEL_EXTENSION}")
        writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
        result_df.to_excel(writer, sheet_name="Summary")
        # And generate a detail sheet, and optionally a pattern sheet and diagrams, for each column
        for column_name, detail_df in detail_dict.items():
            logger.debug(f"Examining column '{column_name}' ...")
            target_sheet_name = make_sheet_name(column_name, MAX_SHEET_NAME_LENGTH-4) + DETAIL_ABBR
            logger.info(f"Writing detail for column '{column_name}' to sheet '{target_sheet_name}' ...")
            detail_df.to_excel(writer, index=False, sheet_name=target_sheet_name)
            if column_name in pattern_dict:
                target_sheet_name = make_sheet_name(column_name, MAX_SHEET_NAME_LENGTH-4) + PATTERN_ABBR
                logger.info(f"Writing pattern information for string column '{column_name}' to sheet '{target_sheet_name}' ...")
                pattern_df = pattern_dict[column_name]
                pattern_df.to_excel(writer, index=False, sheet_name=target_sheet_name)
        writer.close()

        # Add the plots and size bars to the Excel file
        workbook = openpyxl.load_workbook(output_file)

        # Plots
        # Look for sheet names corresponding to the plot filename
        sheet_number = -1
        for sheet_name in workbook.sheetnames:
            sheet_number += 1
            if not sheet_name.endswith(DETAIL_ABBR):
                continue
            if sheet_number == 0:  # Skip summary sheet (first sheet, zero-based-index)
                continue
            column_name = sheet_name[:-len(DETAIL_ABBR)]  # remove " det" from sheet name to get column name
            if column_name in histogram_plot_list:
                sheet_number += 1
                insert_image(HISTOGRAM, sheet_number)
            if column_name in box_plot_list:
                sheet_number += 1
                insert_image(BOX, sheet_number)
            if column_name in pie_plot_list:
                sheet_number += 1
                insert_image(PIE, sheet_number)

        # Formatting for the summary sheet
        worksheet = workbook.worksheets[0]
        worksheet.column_dimensions['A'].width = 25  # Column names
        worksheet.column_dimensions['G'].width = 15  # Most common value
        worksheet.column_dimensions['I'].width = 15  # Largest value
        worksheet.column_dimensions['J'].width = 15  # Smallest value
        worksheet.column_dimensions['K'].width = 15  # Longest value

        for row in range(1, worksheet.max_row+1):
            worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='right')
        for row in range(1, worksheet.max_row+1):
            for col in range(1, 17):
                worksheet.cell(row=row, column=col).border = Border(outline=Side(border_style=borders.BORDER_THICK, color='FFFFFFFF'))

        workbook.save(output_file)
        logger.info(f"Wrote {os.stat(output_file).st_size:,} bytes to '{output_file}'.")


    if is_html_output:
        root_output_dir = tempdir_path / FILE_BASE_NAME
        columns_dir = root_output_dir / "columns"
        images_dir = root_output_dir / "images"
        os.makedirs(columns_dir)
        os.makedirs(images_dir)
        # Move images
        for _, _, files in tempdir_path.walk():
            for file in files:
                os.rename(tempdir_path / file, images_dir / file)
            break
        # Generate a detail page, and optionally a pattern page and diagrams, for each column
        for column_name, detail_df in detail_dict.items():
            logger.debug(f"Examining column '{column_name}' ...")
            target_file = columns_dir / (column_name + DETAIL_ABBR + ".html")
            logger.info(f"Writing detail for column '{column_name}' to '{target_file}' ...")
            with open(target_file, "w") as writer:
                writer.write(make_html_header(f"Exploratory Data Analysis for column: {column_name}", root_output_file=root_output_dir))
                writer.write(f"<h2>Detail analysis for column '{column_name}'</h2>")
                writer.write(f"<hr>")
                writer.write(f"<h3>Value frequency</h3>")
                writer.write(detail_df.to_html(justify="center", na_rep="", index=False))
                if column_name in pattern_dict:
                    logger.info(f"Writing pattern information for string column '{column_name}' to '{target_file}' ...")
                    writer.write(f"<hr>")
                    writer.write(f"<h3>Pattern frequency</h3>")
                    pattern_df = pattern_dict[column_name]
                    writer.write(pattern_df.to_html(justify="center", na_rep="", index=False))
                if column_name in histogram_plot_list:
                    logger.info(f"Adding histogram plot for column '{column_name}' to '{target_file}' ...")
                    writer.write(f"<hr>")
                    writer.write(f"<h3>Histogram</h3>")
                    writer.write(f'<img src="../images/{column_name}.histogram.png" alt="Histogram for column :{column_name}:">')
                if column_name in pie_plot_list:
                    logger.info(f"Adding pie plot for column '{column_name}' to '{target_file}' ...")
                    writer.write(f"<hr>")
                    writer.write(f"<h3>Box plots</h3>")
                    writer.write(f'<img src="../images/{column_name}.pie.png" alt="Pie plot for column :{column_name}:">')
                if column_name in box_plot_list:
                    logger.info(f"Adding box plots for column '{column_name}' to '{target_file}' ...")
                    writer.write(f"<hr>")
                    writer.write(f"<h3>Box plots</h3>")
                    writer.write(f'<img src="../images/{column_name}.box.png" alt="Box plots for column :{column_name}:">')
                writer.write(make_html_footer())
        with open(root_output_dir / f"{FILE_BASE_NAME}.html", "w") as writer:
            logger.info("Writing summary ...")
            writer.write(make_html_header(f"Exploratory Data Analysis for {input}"))
            # Replace column names in summary dataframe with URL links
            replacement_list = [f'<a href="columns/{x} det.html">{x}</a>' for x in result_df.index]
            replacement_dict = dict(zip(result_df.index, replacement_list))
            result_df = result_df.rename(index=replacement_dict)
            writer.write(result_df.to_html(justify="center", na_rep="", escape=False))
            writer.write(make_html_footer())
        logger.info("Making zip archive ...")
        output_file = shutil.make_archive(
            base_name=target_dir / FILE_BASE_NAME,
            format="zip",
            root_dir=tempdir_path,
            base_dir=".",
        )
        logger.info(f"Wrote {os.stat(output_file).st_size:,} bytes to '{output_file}'.")

if cleaned_version_output_file:
    target_path = target_dir / cleaned_version_output_file
    if cleaned_version_output_file.lower().endswith(C.CSV_EXTENSION):
        zipped_target_path = target_path.with_suffix(".zip")
        compression_opts = dict(method='zip', archive_name=cleaned_version_output_file)
        input_df.to_csv(zipped_target_path, index=False, compression=compression_opts)
    else:
        zipped_target_path = target_path.with_suffix(".gz")
        input_df.to_parquet(zipped_target_path, index=False, compression="gzip")
    logger.info(f"Wrote {os.stat(zipped_target_path).st_size:,} bytes to '{zipped_target_path}'.")
